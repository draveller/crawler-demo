import csv
import math
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Any, Union

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from typing_extensions import Literal

# Constants for column/row adjustment
MAX_COLUMN_WIDTH = 40  # Maximum column width in characters
MAX_ROW_HEIGHT = 400  # Maximum row height in points
MAX_IMAGE_WIDTH = 800  # Maximum image width in pixels
MAX_IMAGE_HEIGHT = 600  # Maximum image height in pixels

# 常量定义
BASE_COLUMN_WIDTH = 12  # 基础列宽（英文字符数）
MIN_ROW_HEIGHT = 15  # 最小行高（点）
IMAGE_SCALE_FACTOR = 1  # 图片缩放系数（占单元格比例）


def generate_unique_filename(prefix: str = "data", extension: str = ".xlsx") -> str:
    """生成包含时间戳的唯一文件名"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S%f')
    return f"{prefix}_{timestamp}{extension}"


def get_default_path(store_folder: str = "acer_store") -> str:
    """获取默认存储路径并确保目录存在"""
    current_dir = Path(__file__).resolve().parent
    store_dir = current_dir / store_folder
    store_dir.mkdir(parents=True, exist_ok=True)
    return str(store_dir / generate_unique_filename())


def load_or_create_workbook(excel_path: str) -> Workbook:
    """加载或创建工作簿"""
    try:
        if Path(excel_path).exists():
            try:
                return load_workbook(excel_path)
            except PermissionError as e:
                # 处理文件被占用的情况
                print(f"\n文件 '{excel_path}' 正在被其他程序占用")
                raise SystemExit("程序因文件锁定异常终止") from e
        return Workbook()
    except Exception as e:
        # 处理其他异常
        print(f"\n文件操作失败: {str(e)}, 路径={excel_path}")
        raise

def _append_row_with_images(worksheet: Worksheet, row: List[Any], row_num: int) -> None:
    """追加行数据并处理图片"""
    for col_num, cell_value in enumerate(row, start=1):
        if isinstance(cell_value, Image):
            cell_ref = worksheet.cell(row=row_num, column=col_num).coordinate
            cell_value.anchor = cell_ref
            worksheet.add_image(cell_value)
        else:
            worksheet.cell(row=row_num, column=col_num, value=cell_value)


def _scale_image(image: Image) -> Image:
    """缩放图片保持宽高比"""
    width_ratio = MAX_IMAGE_WIDTH / image.width
    height_ratio = MAX_IMAGE_HEIGHT / image.height
    scale_ratio = min(width_ratio, height_ratio, 1)  # 保持原始比例

    image.width = int(image.width * scale_ratio)
    image.height = int(image.height * scale_ratio)
    return image


def _adjust_columns_rows(worksheet: Worksheet) -> None:
    """智能调整行列尺寸（核心改进）"""
    # ================== 列宽调整 ==================
    # 收集每列的最大有效宽度
    col_metrics = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                # 计算文本有效长度（中文按2字符计算）
                content = str(cell.value)
                effective_length = sum(2 if ord(c) > 127 else 1 for c in content)

                # 根据换行符数量增加宽度补偿
                line_count = max(content.count('\n') + 1, 1)
                adjusted_length = (effective_length / line_count) * 1.2

                # 记录最大值（不超过最大列宽）
                current_max = col_metrics.get(cell.column, {'width': 0, 'line_count': 1})
                new_width = min(max(adjusted_length, current_max['width']), MAX_COLUMN_WIDTH)
                new_line_count = max(line_count, current_max['line_count'])
                col_metrics[cell.column] = {
                    'width': new_width,
                    'line_count': new_line_count
                }

    # 应用列宽调整（考虑基础宽度）
    for col, metrics in col_metrics.items():
        column_letter = get_column_letter(col)
        target_width = max(
            BASE_COLUMN_WIDTH,
            metrics['width'] * (1 + math.log(metrics['line_count'], 5))
        )
        worksheet.column_dimensions[column_letter].width = min(target_width, MAX_COLUMN_WIDTH)

    # ================== 行高调整 ==================
    row_heights = {}
    # 预计算所有图片位置
    image_positions = {}
    for idx, img in enumerate(worksheet._images):
        try:
            # 新版本openpyxl
            img_row = img.anchor._from.row + 1
            img_col = img.anchor._from.col + 1
        except AttributeError:
            # 旧版本openpyxl
            cell_ref = img.anchor
            img_row = int(''.join(filter(str.isdigit, cell_ref)))
            img_col = ord(''.join(filter(str.isalpha, cell_ref)).upper()) - 64

        image_positions[idx] = (img_row, img_col)

    # 遍历所有行
    for row_num, row in enumerate(worksheet.iter_rows(), start=1):
        max_line_height = MIN_ROW_HEIGHT

        # 处理文本行高
        for cell in row:
            if cell.value:
                content = str(cell.value)
                line_count = content.count('\n') + 1
                line_height = (line_count * 14)  # 每行14点
                max_line_height = max(max_line_height, line_height)

        # 处理图片行高
        for img_idx, (img_row, img_col) in image_positions.items():
            if img_row == row_num:
                img = worksheet._images[img_idx]
                # 获取当前列宽
                col_letter = get_column_letter(img_col)
                col_width = worksheet.column_dimensions[col_letter].width

                # 根据列宽计算最大允许尺寸
                max_px_width = (col_width * 7) * IMAGE_SCALE_FACTOR  # 列宽转像素
                max_px_height = (MAX_ROW_HEIGHT * 1.33) * IMAGE_SCALE_FACTOR  # 行高转像素

                # 等比例缩放
                width_ratio = max_px_width / img.width
                height_ratio = max_px_height / img.height
                scale = min(width_ratio, height_ratio, 1)

                img.width = int(img.width * scale)
                img.height = int(img.height * scale)

                # 计算需要的行高
                required_height = max(img.height / 1.33, max_line_height)
                row_heights[row_num] = min(required_height, MAX_ROW_HEIGHT)

        # 设置最终行高
        if row_num in row_heights:
            worksheet.row_dimensions[row_num].height = row_heights[row_num]
        else:
            worksheet.row_dimensions[row_num].height = min(max_line_height, MAX_ROW_HEIGHT)


def save(
        data: Union[List[List[Any]], Tuple[Tuple[Any, ...], ...]],
        excel_path: Union[str, None] = None,
        sheet_name: str = 'Sheet1',
        clean_mode: Literal['wb', 'ws', 'no'] = 'ws'
) -> None:
    """保存数据到Excel文件"""
    excel_path = excel_path or get_default_path()
    wb = load_or_create_workbook(excel_path)

    # 清理策略
    if clean_mode == 'wb':
        for sheet in wb.worksheets[:]:
            wb.remove(sheet)
    elif clean_mode == 'ws' and sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # 获取或创建工作表
    ws = wb.create_sheet(sheet_name) if sheet_name not in wb.sheetnames else wb[sheet_name]

    # 写入数据
    for i, row in enumerate(data, start=1):
        _append_row_with_images(ws, row, i)

    # 调整列宽和行高
    _adjust_columns_rows(ws)

    # 保存工作簿
    try:
        wb.save(excel_path)
    except PermissionError:
        # 抛出异常:
        msg = f"保存失败, 请检查文件 {excel_path} 是否被其他程序打开."
        raise PermissionError(msg)

def save_csv(data: Union[List[List[Any]], Tuple[Tuple[Any, ...], ...]], csv_path: str = None) -> None:
    """保存数据到CSV文件"""
    csv_path = csv_path or "output.csv"
    try:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(data)
        print(f"CSV文件已保存至: {csv_path}")
    except Exception as e:
        print(f"CSV保存失败: {str(e)}")
